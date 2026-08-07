"""
Синхронизация базы контрагентов с Dista Music (вкладка «Dista Connect»).

Наш сервис — мастер всех данных; Dista держит по контрагенту только внутренний
`id` + `Название`. Единственная задача сверки — поддерживать связку
`Contragent.dista_id` и заводить контрагентов, появившихся в Dista.

  POST /dista/reconcile  — загрузка Excel-выгрузки Dista (колонки id + Название),
                           сверка с нашей базой. commit=false → только ПРЕВЬЮ
                           (ничего не пишем); commit=true → применить план.

Сверка ХИРУРГИЧЕСКАЯ и, главное, чего она НЕ делает: на совпавшую карточку
пишет ТОЛЬКО `dista_id`; для новых строк Dista создаёт карточку (название +
dista_id, реквизиты пусты → карточка «неполная»). Бизнес-полей существующих
карточек (номер договора, тип, реквизиты, никнеймы) НЕ касается — в коде нет
присваиваний к ним. Поэтому загрузка id+название ничего не затирает.

Сопоставление: сначала по уже проставленному `dista_id` (железно), затем — для
новых id — по нормализованному имени (наши титлы изначально пришли из этой же
выгрузки Dista, поэтому на старте совпадают). Категории плана:
  link           — имя совпало РОВНО с одной несвязанной карточкой → проставить dista_id
  create         — совпадений нет → создать новую карточку
  already_linked — карточка с этим dista_id уже есть → пропуск (в ответе только счётчик)
  ambiguous      — имя совпало с НЕСКОЛЬКИМИ карточками (напр. аванс/роялти делят
                   титл) → не трогаем, решает человек (связь 1:1)
  only_ours      — наши карточки без dista_id, которых нет в выгрузке → «завести
                   в Dista» (отчёт, ничего не пишем)
  skipped        — строка без id или без названия (связать/создать нечего)
"""
import io
import uuid

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.audit import log_action
from app.auth import get_current_user, require_role
from app.db import get_session
from app.models import Contragent, User
from app.roles import CAN_USE_DISTA_SYNC
from app.tags import build_article

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

dista_router = APIRouter(prefix="/dista", tags=["dista"])

# Заголовки колонок в выгрузке Dista — сопоставляем регистронезависимо. id —
# внутренний идентификатор Dista; название может лежать под «Название»/«Титл».
_ID_HEADERS = {"id", "dista_id", "dista id", "ид"}
_NAME_HEADERS = {"название", "name", "титл", "title", "наименование"}


def _norm(value) -> str:
    """Нормализация имени для сопоставления: схлопнуть пробелы + casefold."""
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _find_col(header: list[str], candidates: set[str]) -> int | None:
    for i, h in enumerate(header):
        if h.strip().casefold() in candidates:
            return i
    return None


def _clean_id(raw) -> str:
    """Приводим id к строке; openpyxl отдаёт целые как float (1183.0) — срезаем '.0'."""
    if raw in (None, ""):
        return ""
    s = str(raw).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


@dista_router.post("/reconcile", dependencies=[Depends(require_role(*CAN_USE_DISTA_SYNC))])
def reconcile(
    file: UploadFile = File(...),
    commit: bool = Form(False),
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Сверка выгрузки Dista с нашей базой. commit=false — превью, commit=true —
    применить (проставить dista_id совпавшим, создать новых). См. докстринг модуля.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Ожидается файл .xlsx")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 — любую ошибку чтения показываем как 400
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}")
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header = [str(v).strip() if v is not None else "" for v in header_row]
    id_idx = _find_col(header, _ID_HEADERS)
    name_idx = _find_col(header, _NAME_HEADERS)
    if id_idx is None or name_idx is None:
        raise HTTPException(
            status_code=400,
            detail="В файле должны быть колонки «id» и «Название» (или «Титл»)",
        )

    # Индексы нашей базы: по dista_id (связанные) и по нормализованному титлу
    # (несвязанные — кандидаты на связывание).
    contragents = db.query(Contragent).all()
    by_dista_id = {c.dista_id: c for c in contragents if c.dista_id}
    unlinked_by_name: dict[str, list[Contragent]] = {}
    for c in contragents:
        if c.dista_id:
            continue
        unlinked_by_name.setdefault(_norm(c.title), []).append(c)

    link: list[dict] = []
    create: list[dict] = []
    ambiguous: list[dict] = []
    skipped: list[dict] = []
    already_linked = 0
    incoming_names: set[str] = set()
    seen_ids: set[str] = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        dista_id = _clean_id(row[id_idx] if id_idx < len(row) else None)
        name = (
            str(row[name_idx]).strip()
            if name_idx < len(row) and row[name_idx] not in (None, "")
            else ""
        )

        if not dista_id:
            skipped.append({"row": row_num, "name": name, "reason": "нет id"})
            continue
        if dista_id in seen_ids:
            skipped.append({"row": row_num, "dista_id": dista_id, "name": name, "reason": "повтор id в файле"})
            continue
        seen_ids.add(dista_id)

        # Уже связано по dista_id — имя не важно, пропускаем.
        if dista_id in by_dista_id:
            already_linked += 1
            incoming_names.add(_norm(by_dista_id[dista_id].title))
            continue

        if not name:
            skipped.append({"row": row_num, "dista_id": dista_id, "reason": "нет названия — карточку не создать"})
            continue
        incoming_names.add(_norm(name))

        candidates = unlinked_by_name.get(_norm(name), [])
        if len(candidates) == 1:
            link.append(
                {"dista_id": dista_id, "name": name, "card_id": str(candidates[0].id), "card_title": candidates[0].title}
            )
        elif len(candidates) > 1:
            ambiguous.append({"dista_id": dista_id, "name": name, "candidates": [c.title for c in candidates]})
        else:
            create.append({"dista_id": dista_id, "name": name})

    only_ours = [
        {"card_id": str(c.id), "title": c.title}
        for c in contragents
        if not c.dista_id and _norm(c.title) not in incoming_names
    ]

    applied = None
    if commit:
        linked = created = 0
        for item in link:
            c = db.get(Contragent, uuid.UUID(item["card_id"]))
            if c is not None and not c.dista_id:
                c.dista_id = item["dista_id"]
                linked += 1
        for item in create:
            db.add(Contragent(title=item["name"], dista_id=item["dista_id"]))
            created += 1
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail="Конфликт при применении (dista_id уже занят). Обновите превью и повторите.",
            )
        log_action(db, current_user, "dista_reconcile", meta={"linked": linked, "created": created})
        applied = {"linked": linked, "created": created}

    return {
        "committed": commit,
        "counts": {
            "link": len(link),
            "create": len(create),
            "already_linked": already_linked,
            "ambiguous": len(ambiguous),
            "only_ours": len(only_ours),
            "skipped": len(skipped),
        },
        "applied": applied,
        "link": link,
        "create": create,
        "ambiguous": ambiguous,
        "only_ours": only_ours,
        "skipped": skipped,
    }


@dista_router.get("/status", dependencies=[Depends(require_role(*CAN_USE_DISTA_SYNC))])
def status(db: Session = Depends(get_session)) -> dict:
    """Сводка связки для страницы: сколько карточек связано с Dista, сколько нет."""
    total = db.query(Contragent).count()
    linked = db.query(Contragent).filter(Contragent.dista_id.isnot(None)).count()
    return {"total": total, "linked": linked, "unlinked": total - linked}


@dista_router.get("/only-ours-export", dependencies=[Depends(require_role(*CAN_USE_DISTA_SYNC))])
def only_ours_export(
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    """
    Выгрузка «Нет в Dista» (Этап 2): наши карточки без dista_id — те, кого нужно
    завести в Dista вручную (импорта туда пока нет). Колонки: Название +
    Артикул (+ контекст). Артикул вписывают в заметку карточки в Dista — тогда
    следующая сверка свяжет их по нему автоматически (Этап 3), а не по имени.
    """
    rows = (
        db.query(Contragent)
        .filter(Contragent.dista_id.is_(None))
        .order_by(Contragent.title)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Нет в Dista"
    ws.append(["Название", "Артикул", "Страна", "Тип", "Рег. номер"])
    for c in rows:
        ws.append([
            c.title,
            build_article(c.country, c.reg_number) or "",
            c.country or "",
            c.type or "",
            c.reg_number or "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action(db, current_user, "dista_only_ours_export", meta={"rows": len(rows)})

    return StreamingResponse(
        buffer,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="dista_to_add.xlsx"'},
    )
