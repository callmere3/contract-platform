"""
Эндпоинты работы с контрагентами (этап 4, база контрагентов).

  POST /contragents                    — создать контрагента (title/contract_number
                                          вычисляются автоматически, см. брейншторм)
  GET  /contragents?q=...              — поиск по title/nickname (ILIKE, регистронезависимо).
                                          Без q — весь список (для стартового экрана/отладки).
  POST /contragents/import             — массовый импорт из .xlsx (обязательна только
                                          колонка "Титл"; дубли по title обновляются,
                                          а не дублируются)
  GET  /contragents/export             — выгрузка всех контрагентов в .xlsx (тот же
                                          формат колонок, что и импорт — файл можно
                                          поправить руками и залить обратно)
  GET  /contragents/{id}               — карточка контрагента целиком
  PATCH /contragents/{id}              — правка карточки (кроме title; пока
                                          только через Swagger, без UI)
  DELETE /contragents/{id}             — удалить контрагента (никнеймы каскадно;
                                          пока только через Swagger, без кнопки в UI)
  GET  /contragents/{id}/templates     — подбор документов, совместимых с контрагентом
                                          по тегам (country/contragent_type/contract_family)
  POST /contragents/{id}/nicknames     — добавить псевдоним контрагенту

ВАЖНО про порядок маршрутов: /import и /export зарегистрированы РАНЬШЕ
/{contragent_id} — иначе FastAPI попытался бы распарсить 'import'/'export'
как uuid.UUID и вернул бы 422 вместо реального обработчика (порядок
регистрации маршрутов в Starlette имеет значение для статических путей
против путей с параметром).
"""
import io
import json
import uuid
from datetime import date as _date
from datetime import datetime as _datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, case, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.audit import log_action
from app.auth import get_current_user, require_role
from app.context_builder import build_contract_number, build_contragent_title, parse_date
from app.db import get_session
from app.models import Contragent, ContragentNickname, Template, User, doc_type_sort_key
from app.roles import (
    ADMIN,
    CAN_CREATE_CONTRAGENTS,
    CAN_EDIT_CONTRACT_FAMILY,
    CAN_EDIT_CONTRAGENTS,
    CAN_EDIT_REG_NUMBER,
    CAN_EDIT_REQUISITES,
    CAN_EXPORT_CONTRAGENTS,
    SEES_HIDDEN_TEMPLATES,
)
from app.tags import (
    ALL_REQUISITE_FIELDS,
    COUNTRIES,
    CONTRACT_FAMILIES,
    CONTRAGENT_TYPES,
    OBLIGATION_DOC_TYPES,
    REQUISITE_FIELDS_BY_TYPE,
    normalize_optional_tag,
    normalize_reg_number,
    normalize_tag,
    obligation_bucket,
)

contragents_router = APIRouter(prefix="/contragents", tags=["contragents"])

# Порядок колонок общий для импорта и экспорта (см. брейншторм, раздел
# "Импорт/экспорт (Excel)") — человекочитаемые русские заголовки, файл
# симметричен в обе стороны: экспортировал -> поправил руками -> залил
# обратно тем же импортом.
#
# "Титл" — единственная обязательная колонка (это title контрагента,
# по нему же идёт поиск/дедупликация при импорте). "Название" (ФИО/
# название компании) — необязательная колонка, в отличие от создания
# через UI (POST /contragents), где name обязателен, а title вычисляется
# из него автоматически. При импорте — наоборот: title всегда берётся из
# файла как есть, а name опционален.
EXCEL_COLUMNS = [
    "Титл", "Название", "Никнеймы", "Тип", "Страна",
    "Тип договора", "Номер договора", "Дата договора", "Роялти %",
    "Рег. номер",
]
# "Рег. номер" — ИНН (ФЛ/СГ) / ОГРНИП (ИП) / ОГРН (ООО) / БИН (ТОО), см.
# app/tags.py: REG_NUMBER_META. Одна колонка на все смыслы, как и в самой БД.


def _contragent_is_complete(c: Contragent) -> bool:
    """
    Карточка заполнена ПОЛНОСТЬЮ = заполнены поля, которые менеджер видит в
    карточке контрагента (ContragentCardModal.ROWS на фронте): ФИО, страна,
    тип, рег.номер, номер и дата договора, роялти.

    Псевдоним НЕ входит: он тоже показывается в карточке, но легально бывает
    пустым ("если нет — оставьте пустым") — согласовано с пользователем.
    reg_number, наоборот, входит, хоть и необязателен при создании: карточка
    без ИНН/ОГРН считается неполной (это сигнал дозаполнить).

    contract_family (тип договора) в проверку НЕ входит (убрано по просьбе
    пользователя): контрагент не привязан к одному семейству договора — по
    стране и типу ему подходят сразу все (аванс/роялти/аванс+обязательство,
    см. list_contragent_templates). Поэтому пустой contract_family больше не
    повод красить карточку, а сам подбор документов на него не смотрит.

    Используется для подсветки неполных карточек в "Базе контрагентов"
    (ContragentRow) — контрагент может быть заведён "неполным" через импорт
    (большинство полей nullable, см. докстринг Contragent). royalty_percent
    и contract_date проверяем на None явно: 0% роялти — валидное значение,
    а не "пусто".
    """
    return all(
        [
            c.name,
            c.country,
            c.type,
            c.reg_number,
            c.contract_number,
            c.contract_date is not None,
            c.royalty_percent is not None,
        ]
    )


def _contragent_summary(c: Contragent) -> dict:
    """Краткое представление для списков поиска — title крупно, никнеймы мелко (см. брейншторм)."""
    return {
        "id": str(c.id),
        "title": c.title,
        "name": c.name,
        "nicknames": [n.nickname for n in c.nicknames],
        "country": c.country,
        "type": c.type,
        "contract_family": c.contract_family,
        "reg_number": c.reg_number,
        # флаг для подсветки неполных карточек в списке (см. _contragent_is_complete)
        "is_complete": _contragent_is_complete(c),
    }


def _parse_requisites(raw: str | None, contragent_type: str | None) -> dict | None:
    """
    Разбирает реквизиты карточки из JSON-строки формы (Contragent.requisites).

    Форма шлёт реквизиты одним JSON-объектом {имя_метки: значение} — так их
    удобно принять в общем Form-эндпоинте рядом с остальными полями, не
    заводя колонку на каждый реквизит.

    Оставляем только КЛЮЧИ, допустимые для этого типа контрагента
    (REQUISITE_FIELDS_BY_TYPE) — чужие/незнакомые ключи отбрасываем, чтобы в
    карточку не попадал мусор или поля не того типа. Тип неизвестен -> любой
    известный реквизит (ALL_REQUISITE_FIELDS) как мягкий фолбэк. Пустые
    значения выкидываем (незаполненное поле = «нет реквизита»), значения
    приводим к строке и обрезаем пробелы.

    None -> None (поле не передано, «не трогать» — см. update_contragent).
    Пустой после чистки словарь -> {} (реквизиты явно очищены).
    """
    if raw is None:
        return None
    try:
        data = json.loads(raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Реквизиты должны быть корректным JSON-объектом")
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="Реквизиты должны быть объектом {поле: значение}")

    allowed = set(REQUISITE_FIELDS_BY_TYPE.get(contragent_type or "", ())) or ALL_REQUISITE_FIELDS
    cleaned: dict[str, str] = {}
    for key, value in data.items():
        if key not in allowed:
            continue
        text = ("" if value is None else str(value)).strip()
        if text:
            cleaned[key] = text
    return cleaned


def _parse_excel_date(value) -> _date | None:
    """
    Ячейка "Дата договора" — Excel сам отдаёт datetime/date для ячеек с
    форматом даты (openpyxl, data_only=True), но допускаем и текст на
    случай, если колонку сохранили как обычный текст — тогда используем
    тот же parse_date(), что и в остальном проекте (ISO/русский/точечный
    форматы), чтобы не заводить третий парсер дат в одном сервисе.
    """
    if value in (None, ""):
        return None
    if isinstance(value, _datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    parsed = parse_date(str(value))
    if not parsed:
        return None
    day, month, year = parsed
    return _date(int(year), int(month), int(day))


def _parse_percent(value) -> Decimal | None:
    """Ячейка "Роялти %" — число или текст с запятой вместо точки ('12,5')."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", ".").replace("%", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _try_normalize_optional(value, allowed: list[str], field_name: str) -> tuple[str | None, str | None]:
    """
    Как normalize_optional_tag, но не бросает исключение — возвращает
    (значение_или_None, текст_предупреждения_или_None). Импорт работает
    "мягко" (см. брейншторм — "validate loosely, не all-or-nothing"):
    опечатка в одной ячейке одной строки не должна ронять всю загрузку
    файла на сотни контрагентов — теряется только это поле этой строки,
    остальное (включая title/nickname) сохраняется, а причина попадает
    в отчёт по этой строке.
    """
    if value in (None, ""):
        return None, None
    try:
        return normalize_tag(str(value), allowed, field_name), None
    except HTTPException as exc:
        return None, str(exc.detail)


def _try_normalize_reg_number(value, contragent_type: str | None) -> tuple[str | None, str | None]:
    """Как normalize_reg_number(), но мягко — см. _try_normalize_optional() выше."""
    if value in (None, ""):
        return None, None
    try:
        return normalize_reg_number(str(value), contragent_type), None
    except HTTPException as exc:
        return None, str(exc.detail)


def _filtered_contragents_query(
    db: Session,
    q: str | None = None,
    country: str | None = None,
    contragent_type: str | None = None,
):
    """
    Общий запрос контрагентов с фильтрами q/country/contragent_type —
    единый источник для поиска (search_contragents) и экспорта
    (export_contragents), чтобы экспорт выгружал ровно то же, что показано
    в списке при текущем фильтре. Без .limit()/.offset() — пагинацию
    навешивает вызывающий, экспорт выгружает всё.

    Поиск по title, name ИЛИ nickname (см. брейншторм): не важно, по чему
    совпало. country/contragent_type — точное совпадение после нормализации
    регистра (app/tags.py), применяются ДОПОЛНИТЕЛЬНО к q.

    Сортировка при поиске: сначала совпадения "с начала строки", потом
    подстрочные, внутри группы — по алфавиту (по просьбе пользователя). Ввёл
    "иба" — "Ибара" (title с начала) идёт раньше "Дарибаева" (иба в середине).
    Ранжирование делается в SQL, а не в Python: список постраничный, и
    префиксное совпадение с дальней страницы иначе не поднялось бы наверх.

    Совпадение по nickname проверяется через EXISTS, а не outerjoin+distinct:
    у контрагента несколько псевдонимов, join размножил бы строки, а с
    ORDER BY по вычисляемому рангу distinct на джойне становится неудобным.
    EXISTS даёт по одной строке на контрагента без дублей.
    """
    query = db.query(Contragent)
    order_by = [Contragent.title]
    if q:
        like = f"%{q}%"
        prefix = f"{q}%"
        nickname_matches = (
            db.query(ContragentNickname.id)
            .filter(
                ContragentNickname.contragent_id == Contragent.id,
                ContragentNickname.nickname.ilike(like),
            )
            .exists()
        )
        nickname_prefix = (
            db.query(ContragentNickname.id)
            .filter(
                ContragentNickname.contragent_id == Contragent.id,
                ContragentNickname.nickname.ilike(prefix),
            )
            .exists()
        )
        query = query.filter(
            or_(
                Contragent.title.ilike(like),
                Contragent.name.ilike(like),
                nickname_matches,
            )
        )
        starts_with = or_(
            Contragent.title.ilike(prefix),
            Contragent.name.ilike(prefix),
            nickname_prefix,
        )
        order_by = [case((starts_with, 0), else_=1), Contragent.title]
    if country:
        query = query.filter(
            Contragent.country == normalize_optional_tag(country, COUNTRIES, "country")
        )
    if contragent_type:
        query = query.filter(
            Contragent.type
            == normalize_optional_tag(contragent_type, CONTRAGENT_TYPES, "contragent_type")
        )
    return query.order_by(*order_by)


@contragents_router.get("", dependencies=[Depends(get_current_user)])
def search_contragents(
    q: str | None = None,
    country: str | None = None,
    contragent_type: str | None = None,
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_session),
) -> dict:
    """
    Поиск по title, name ИЛИ nickname одновременно (см. брейншторм — SQL-
    запрос там же): не важно, по чему совпало, оператор в любом случае
    видит и title, и никнеймы сразу. name добавлен отдельно от title,
    потому что это разные строки (title — вычисленное "Иванов И. И. (СГ)",
    name — сырое "Иванов Иван Иванович") — поиск/проверка дублей по одному
    только title могла не найти уже существующего человека, если искать
    по тому, как его ФИО выглядит целиком.

    country/contragent_type — необязательные фильтры по тегам (точное
    совпадение после нормализации регистра, см. app/tags.py), применяются
    ДОПОЛНИТЕЛЬНО к q, а не вместо него. contract_family сюда намеренно
    не добавлен — фильтр по роялти/авансу не нужен (см. запрос пользователя).

    Пагинация: page (с 1), page_size (по умолчанию 100, максимум 500). Раньше
    список резался жёстким лимитом 200, и при большой базе (800+ контрагентов
    после боевого импорта) остальных было не посмотреть. Теперь в ответе есть
    total — общее число записей под текущим фильтром, чтобы фронт нарисовал
    "показаны X–Y из N" и листалку. Экспорт по-прежнему выгружает всё без
    пагинации (см. export_contragents).
    """
    page = max(1, page)
    page_size = max(1, min(page_size, 500))
    base = _filtered_contragents_query(db, q, country, contragent_type)
    total = base.count()
    rows = base.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "contragents": [_contragent_summary(c) for c in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@contragents_router.post("", dependencies=[Depends(require_role(*CAN_CREATE_CONTRAGENTS))])
def create_contragent(
    name: str = Form(...),
    country: str = Form(...),           # 'РУ' | 'КЗ'
    contragent_type: str = Form(...),   # 'ФЛ' | 'СГ' | 'ИП' | 'ООО' | 'ТОО'
    contract_family: str = Form(...),   # 'РОЯЛТИ' | 'АВАНС' | 'АВАНС_ОБЯЗАТЕЛЬСТВО'
    contract_date: str = Form(...),     # ISO из <input type="date">, напр. '2026-03-15'
    royalty_percent: float = Form(...),
    reg_number: str | None = Form(None),  # ИНН (СГ) / ОГРНИП (ИП) / ОГРН (ООО)
    nicknames: str | None = Form(None), # через запятую, тот же формат, что и в импорте
    requisites: str | None = Form(None),  # JSON {имя_метки: значение}, необязательно
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Создаёт карточку контрагента. title и contract_number вычисляются
    автоматически и НЕ принимаются от клиента — см. брейншторм
    ("title... НЕ показывается и НЕ редактируется в форме создания").

    reg_number — необязателен (контрагент может быть создан "неполным"),
    но если передан — проверяется на цифры и длину под contragent_type
    (см. app/tags.py: REG_NUMBER_META) и на уникальность: карточка с таким
    же рег. номером уже существующего контрагента не создастся — это и
    есть точный идентификатор, а не мягкая проверка по title (см. ниже).

    contract_date фиксируется здесь один раз и дальше только отображается
    при генерации "Договора" — не пересчитывается на лету (см. брейншторм,
    "Почему именно так, а не иначе").

    nicknames — необязательное поле, через запятую (как в импорте/экспорте,
    см. import_contragents), можно оставить пустым и добавить никнейм(ы)
    позже через POST /contragents/{id}/nicknames.

    Мягкая проверка дублей (см. брейншторм) — задача фронтенда: перед
    сабмитом вызвать GET /contragents?q=<name> и показать похожие; сам
    этот эндпоинт ничего не блокирует (жёсткого constraint на title в БД
    нет, см. models.py).

    Это эндпоинт СОЗДАНИЯ ЧЕРЕЗ UI — остальные поля обязательны, как в
    UX-флоу брейншторма ("Новый контрагент"). Массовый импорт с неполными
    записями (только title/nickname) — отдельный эндпоинт (POST
    /contragents/import, шаг 5, работает по другим правилам: title
    берётся из файла как есть, без пересчёта).

    country/contragent_type/contract_family нормализуются к каноническому
    регистру (см. app/tags.py) — без этого 'ру' и 'РУ' в БД считались бы
    разными значениями, и подбор документов по тегам (шаг 4) молча не
    находил бы шаблоны для контрагента, введённого не в том регистре.
    """
    country = normalize_tag(country, COUNTRIES, "country")
    contragent_type = normalize_tag(contragent_type, CONTRAGENT_TYPES, "contragent_type")
    contract_family = normalize_tag(contract_family, CONTRACT_FAMILIES, "contract_family")

    if not (0 <= royalty_percent <= 100):
        raise HTTPException(
            status_code=400,
            detail=f"Роялти должно быть числом от 0 до 100, получено: {royalty_percent}",
        )

    parsed = parse_date(contract_date)
    if not parsed:
        raise HTTPException(
            status_code=400, detail=f"Не удалось распознать дату договора: {contract_date!r}"
        )
    day, month, year_full = parsed

    reg_number = normalize_reg_number(reg_number, contragent_type)
    if reg_number is not None:
        existing_by_reg = (
            db.query(Contragent).filter(Contragent.reg_number == reg_number).one_or_none()
        )
        if existing_by_reg is not None:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Контрагент с рег. номером {reg_number!r} уже существует: "
                    f"{existing_by_reg.title!r}"
                ),
            )

    title = build_contragent_title(name, contragent_type)
    contract_number = build_contract_number(
        day=day,
        month=month,
        year=year_full[-2:],
        full_name=name,
        doc_kind=contragent_type,
    )

    # Реквизиты при создании необязательны (по решению владельца) — пустой
    # блок допустим, дозаполнят позже в карточке. Чужие/неизвестные ключи и
    # пустые значения отсекаются (_parse_requisites). None/{} -> None в БД.
    requisites_dict = _parse_requisites(requisites, contragent_type) or None

    contragent = Contragent(
        name=name,
        title=title,
        country=country,
        type=contragent_type,
        contract_family=contract_family,
        contract_date=_date(int(year_full), int(month), int(day)),
        contract_number=contract_number,
        royalty_percent=royalty_percent,
        reg_number=reg_number,
        requisites=requisites_dict,
    )
    db.add(contragent)
    db.flush()   # нужен contragent.id до вставки никнеймов

    nickname_list = (
        [n.strip() for n in nicknames.split(",") if n.strip()] if nicknames else []
    )
    for nick in nickname_list:
        db.add(ContragentNickname(contragent_id=contragent.id, nickname=nick))

    try:
        db.commit()
    except IntegrityError:
        # страховка на случай гонки: явную проверку выше кто-то мог
        # обогнать между SELECT и INSERT — БД всё равно не даст создать
        # дубль по уникальному индексу reg_number.
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Контрагент с рег. номером {reg_number!r} уже существует",
        )

    log_action(
        db, current_user, "contragent.create", entity_type="contragent", entity_id=contragent.id,
        meta={"title": contragent.title},
    )

    return {
        "id": str(contragent.id),
        "title": contragent.title,
        "contract_number": contragent.contract_number,
        "contract_date": contragent.contract_date.isoformat(),
        "reg_number": contragent.reg_number,
        "nicknames": nickname_list,
    }


@contragents_router.post("/import", dependencies=[Depends(require_role(ADMIN))])
def import_contragents(
    file: UploadFile = File(...),
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Массовый импорт контрагентов из .xlsx — одна строка на контрагента
    (см. брейншторм, раздел "Импорт/экспорт (Excel)").

    Обязательна ТОЛЬКО колонка "Титл" (title) — по ней же идёт поиск
    дублей. Остальные колонки, включая "Название" (ФИО/название), могут
    отсутствовать или быть пустыми в конкретной строке, контрагент
    создаётся/обновляется "неполным" (см. GET /contragents/{id}/templates —
    такой контрагент просто не участвует в подборе документов, пока
    карточку не дозаполнят).

    "Титл" и "Номер договора" берутся из файла КАК ЕСТЬ, без пересчёта —
    в отличие от POST /contragents (создание через UI), где title
    вычисляется из name по формуле. Здесь это исторические/юридически
    зафиксированные значения, пересчёт мог бы дать другое число, чем
    реально стоит в бумажном договоре.

    Дубли — точное совпадение "Титл" с уже существующим контрагентом:
    существующая запись ОБНОВЛЯЕТСЯ, вторая не создаётся. При обновлении
    непустая ячейка перезаписывает соответствующее поле, пустая — оставляет
    прежнее значение как есть (не затирает то, что уже было заполнено
    вручную и отсутствует в конкретном файле повторного импорта). Никнеймы
    при непустой ячейке заменяются ПОЛНОСТЬЮ новым набором из ячейки
    (через запятую), а не дополняются.

    Невалидный тег (country/тип/тип договора — опечатка, значения нет
    среди допустимых) не роняет всю строку: поле остаётся пустым, причина
    попадает в "details" по этой строке, остальные поля сохраняются.
    Единственная причина полностью пропустить строку — отсутствие "Титл".
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Ожидается файл .xlsx")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}")
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header = [str(v).strip() if v is not None else "" for v in header_row]
    if "Титл" not in header:
        raise HTTPException(
            status_code=400, detail="В файле нет обязательной колонки «Титл»"
        )
    col_index = {name: i for i, name in enumerate(header)}

    def cell(row: tuple, name: str):
        idx = col_index.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    created = updated = skipped = 0
    details: list[dict] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue   # полностью пустая строка — не считаем ни пропуском, ни ошибкой

        title_raw = cell(row, "Титл")
        title = str(title_raw).strip() if title_raw not in (None, "") else ""
        if not title:
            skipped += 1
            details.append({"row": row_num, "status": "пропущено", "reason": "нет титла (title)"})
            continue

        name_raw = cell(row, "Название")
        name = str(name_raw).strip() if name_raw not in (None, "") else None

        warnings: list[str] = []

        contragent_type, w = _try_normalize_optional(cell(row, "Тип"), CONTRAGENT_TYPES, "Тип")
        if w:
            warnings.append(w)
        country, w = _try_normalize_optional(cell(row, "Страна"), COUNTRIES, "Страна")
        if w:
            warnings.append(w)
        contract_family, w = _try_normalize_optional(
            cell(row, "Тип договора"), CONTRACT_FAMILIES, "Тип договора"
        )
        if w:
            warnings.append(w)

        contract_number_raw = cell(row, "Номер договора")
        contract_number = (
            str(contract_number_raw).strip() if contract_number_raw not in (None, "") else None
        )
        contract_date_val = _parse_excel_date(cell(row, "Дата договора"))
        royalty_percent = _parse_percent(cell(row, "Роялти %"))

        nicknames_raw = cell(row, "Никнеймы")
        nicknames = (
            [n.strip() for n in str(nicknames_raw).split(",") if n.strip()]
            if nicknames_raw not in (None, "")
            else None
        )

        existing = db.query(Contragent).filter(Contragent.title == title).one_or_none()

        # тип для проверки длины reg_number — из этой же строки, а если
        # там пусто (ячейка "Тип" не заполнена в файле) — из уже
        # существующей записи (обновление без изменения типа).
        reg_number_type = contragent_type or (existing.type if existing else None)
        reg_number, w = _try_normalize_reg_number(cell(row, "Рег. номер"), reg_number_type)
        if w:
            warnings.append(w)
        elif reg_number is not None:
            conflict = (
                db.query(Contragent)
                .filter(
                    Contragent.reg_number == reg_number,
                    Contragent.id != (existing.id if existing else None),
                )
                .one_or_none()
            )
            if conflict is not None:
                warnings.append(
                    f"рег. номер {reg_number!r} уже занят контрагентом {conflict.title!r} — "
                    f"не записан для этой строки"
                )
                reg_number = None

        if existing is None:
            contragent = Contragent(
                name=name,   # теперь отдельная опциональная колонка "Название", не заглушка
                title=title,
                country=country,
                type=contragent_type,
                contract_family=contract_family,
                contract_date=contract_date_val,
                contract_number=contract_number,
                royalty_percent=royalty_percent,
                reg_number=reg_number,
            )
            db.add(contragent)
            db.flush()   # нужен contragent.id до вставки никнеймов
            for nick in (nicknames or []):
                db.add(ContragentNickname(contragent_id=contragent.id, nickname=nick))
            created += 1
            details.append(
                {"row": row_num, "status": "создано", "title": title, "warnings": warnings}
            )
        else:
            if name is not None:
                existing.name = name
            if country is not None:
                existing.country = country
            if contragent_type is not None:
                existing.type = contragent_type
            if contract_family is not None:
                existing.contract_family = contract_family
            if contract_number is not None:
                existing.contract_number = contract_number
            if contract_date_val is not None:
                existing.contract_date = contract_date_val
            if royalty_percent is not None:
                existing.royalty_percent = royalty_percent
            if reg_number is not None:
                existing.reg_number = reg_number
            if nicknames is not None:
                for old_nick in list(existing.nicknames):
                    db.delete(old_nick)
                db.flush()
                for nick in nicknames:
                    db.add(ContragentNickname(contragent_id=existing.id, nickname=nick))
            updated += 1
            details.append(
                {"row": row_num, "status": "обновлено", "title": title, "warnings": warnings}
            )

    try:
        db.commit()
    except IntegrityError:
        # страховка на случай гонки с параллельным запросом (см. create_contragent) —
        # проверки выше делались построчно в рамках этой же транзакции и
        # не видят изменений из ДРУГИХ, ещё не закоммиченных транзакций.
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=(
                "Не удалось сохранить импорт: конфликт по уникальному рег. номеру "
                "с записью, созданной параллельно. Повторите импорт."
            ),
        )

    log_action(
        db, current_user, "contragent.import", entity_type="contragent",
        meta={"created": created, "updated": updated, "skipped": skipped},
    )

    return {"created": created, "updated": updated, "skipped": skipped, "details": details}


@contragents_router.get("/export", dependencies=[Depends(require_role(*CAN_EXPORT_CONTRAGENTS))])
def export_contragents(
    q: str | None = None,
    country: str | None = None,
    contragent_type: str | None = None,
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    """
    Выгружает контрагентов в .xlsx — те же колонки и тот же порядок, что
    ожидает import_contragents(), файл можно поправить руками и залить
    обратно тем же импортом (см. брейншторм — "Экспорт... симметрично
    формату импорта").

    q/country/contragent_type — те же фильтры, что и у поиска
    (_filtered_contragents_query): экспорт выгружает ровно то, что видно в
    "Базе контрагентов" при текущем фильтре (напр. только РУ, или только
    КЗ+ИП). Без фильтров — вся база. В отличие от поиска, без лимита 200:
    выгружаем всё, что попало под фильтр.

    Доступно Admin/Director/TopManager (см. CAN_EXPORT_CONTRAGENTS) —
    выгрузка базы наружу считается более "чувствительным" действием, чем
    обычное создание/редактирование карточки, поэтому не дана Manager.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контрагенты"
    ws.append(EXCEL_COLUMNS)

    contragents = _filtered_contragents_query(db, q, country, contragent_type).all()
    for c in contragents:
        ws.append([
            c.title,
            c.name or "",
            ", ".join(n.nickname for n in c.nicknames),
            c.type or "",
            c.country or "",
            c.contract_family or "",
            c.contract_number or "",
            c.contract_date.isoformat() if c.contract_date else "",
            float(c.royalty_percent) if c.royalty_percent is not None else "",
            c.reg_number or "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action(
        db, current_user, "contragent.export", entity_type="contragent",
        meta={"rows": len(contragents)},
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="contragents_export.xlsx"'},
    )


@contragents_router.get("/{contragent_id}", dependencies=[Depends(get_current_user)])
def get_contragent(contragent_id: uuid.UUID, db: Session = Depends(get_session)) -> dict:
    """Полная карточка контрагента — для экрана "после выбора контрагента" (см. брейншторм)."""
    contragent = db.get(Contragent, contragent_id)
    if contragent is None:
        raise HTTPException(status_code=404, detail="Контрагент не найден")

    return {
        "id": str(contragent.id),
        "name": contragent.name,
        "title": contragent.title,
        "country": contragent.country,
        "type": contragent.type,
        "contract_family": contragent.contract_family,
        "contract_date": (
            contragent.contract_date.isoformat() if contragent.contract_date else None
        ),
        "contract_number": contragent.contract_number,
        "royalty_percent": (
            float(contragent.royalty_percent) if contragent.royalty_percent is not None else None
        ),
        "reg_number": contragent.reg_number,
        "nicknames": [n.nickname for n in contragent.nicknames],
        # Реквизиты (адреса, банк, паспорт СГ…) — словарь {имя_метки: значение};
        # пусто -> {} (фронт рисует сворачиваемый блок из /tags-описаний).
        "requisites": contragent.requisites or {},
    }


@contragents_router.patch("/{contragent_id}", dependencies=[Depends(require_role(*CAN_EDIT_CONTRACT_FAMILY))])
def update_contragent(
    contragent_id: uuid.UUID,
    title: str | None = Form(None),
    name: str | None = Form(None),
    country: str | None = Form(None),
    contragent_type: str | None = Form(None),
    contract_family: str | None = Form(None),
    contract_date: str | None = Form(None),
    contract_number: str | None = Form(None),
    royalty_percent: str | None = Form(None),
    reg_number: str | None = Form(None),
    nicknames: str | None = Form(None),
    requisites: str | None = Form(None),  # JSON {имя_метки: значение}, полная замена
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Правит существующую карточку контрагента.

    title и contract_number ЗДЕСЬ НЕ ПЕРЕСЧИТЫВАЮТСЯ (осознанное решение
    31.07.2026, разворот прежнего от 16.07.2026). Причина: все титлы
    соответствуют существующей базе компании, а старые номера договоров
    собраны по ДРУГОЙ формуле, чем build_contract_number() — любой пересчёт
    задним числом разошёлся бы с уже подписанными документами. Поэтому
    правка ФИО/типа/даты карточки НЕ трогает ни подпись, ни номер договора.
    Массово их задаёт только импорт из Excel; здесь остаётся лишь ручной
    override — явно переданное непустое значение title/contract_number
    сохраняется как есть (title пустым быть не может — NOT NULL, вернём 400).

    Ручной override title нужен под реальный кейс: один человек может иметь
    ДВЕ карточки (contract_family=АВАНС и РОЯЛТИ) — при создании обе получают
    ОДИНАКОВЫЙ title (формула зависит только от name+type), их разделяют
    вручную ("Иванов И. И. (СГ, аванс)" / "(СГ, роялти)"), иначе не различить
    в поиске. На фронте формы для title/номера нет — через UI эти поля не
    редактируются вовсе (см. EditContragentModal).

    Семантика остальных полей — та же, что и в PATCH /templates/{id}:
      - параметр не передан в форме -> не трогаем, значение остаётся как было
      - передана пустая строка -> поле очищается (None)
      - передано непустое значение -> валидируется/парсится и сохраняется

    nicknames — при непустом значении ПОЛНОСТЬЮ заменяет прежний список
    (через запятую), как и при импорте, а не дополняет его.

    reg_number валидируется под ТЕКУЩИЙ тип контрагента: если contragent_type
    передан этим же запросом — под новый, иначе под уже сохранённый в
    карточке. Проверка уникальности исключает саму карточку (можно сохранить
    то же значение повторно, это не конфликт с самим собой).
    """
    contragent = db.get(Contragent, contragent_id)
    if contragent is None:
        raise HTTPException(status_code=404, detail="Контрагент не найден")

    # Роли вне CAN_EDIT_CONTRAGENTS (сейчас — manager) правят тип договора,
    # реквизиты И рег. номер (все три открыты всем ролям по решению владельца —
    # CAN_EDIT_CONTRACT_FAMILY / CAN_EDIT_REQUISITES / CAN_EDIT_REG_NUMBER).
    # Любое ДРУГОЕ переданное поле — 403. Это серверная защита, не только UI:
    # модалка менеджеру показывает ограниченный набор, но запрос можно подделать.
    # contract_family / requisites / reg_number в список ниже НЕ входят.
    if current_user.role not in CAN_EDIT_CONTRAGENTS:
        others = (
            title, name, country, contragent_type, contract_date,
            contract_number, royalty_percent, nicknames,
        )
        if any(v is not None for v in others):
            raise HTTPException(
                status_code=403,
                detail="Вам доступно изменение только типа договора, реквизитов и рег. номера контрагента",
            )
        if requisites is not None and current_user.role not in CAN_EDIT_REQUISITES:
            raise HTTPException(status_code=403, detail="Правка реквизитов недоступна")
        if reg_number is not None and current_user.role not in CAN_EDIT_REG_NUMBER:
            raise HTTPException(status_code=403, detail="Правка рег. номера недоступна")

    if name is not None:
        contragent.name = name.strip() or None

    if country is not None:
        contragent.country = normalize_optional_tag(country, COUNTRIES, "country")
    if contragent_type is not None:
        contragent.type = normalize_optional_tag(
            contragent_type, CONTRAGENT_TYPES, "contragent_type"
        )
    if contract_family is not None:
        contragent.contract_family = normalize_optional_tag(
            contract_family, CONTRACT_FAMILIES, "contract_family"
        )

    if contract_date is not None:
        if not contract_date.strip():
            contragent.contract_date = None
        else:
            parsed = parse_date(contract_date)
            if not parsed:
                raise HTTPException(
                    status_code=400,
                    detail=f"Не удалось распознать дату договора: {contract_date!r}",
                )
            day, month, year_full = parsed
            contragent.contract_date = _date(int(year_full), int(month), int(day))

    # ---- title и contract_number: ПРИ ПРАВКЕ КАРТОЧКИ НЕ ПЕРЕСЧИТЫВАЮТСЯ ----
    # Осознанное решение 31.07.2026 (разворот прежнего от 16.07.2026): все
    # титлы соответствуют существующей базе компании, а старые номера
    # договоров собраны по ДРУГОЙ формуле, чем build_contract_number() —
    # пересчёт задним числом разошёлся бы с уже подписанными документами.
    # Поэтому правка ФИО/типа/даты карточки НЕ трогает ни подпись, ни номер.
    # Остаётся только ручной override — явно переданное непустое значение
    # сохраняется как есть; массово title/номер задаёт лишь импорт из Excel.
    # (На фронте формы для этих полей нет, см. EditContragentModal — через
    # UI они не редактируются вовсе, только приезжают импортом.)
    if title is not None:
        stripped_title = title.strip()
        if not stripped_title:
            raise HTTPException(status_code=400, detail="title не может быть пустым")
        contragent.title = stripped_title

    if contract_number is not None:
        contragent.contract_number = contract_number.strip() or None

    if royalty_percent is not None:
        if not royalty_percent.strip():
            contragent.royalty_percent = None
        else:
            value = _parse_percent(royalty_percent)
            if value is None:
                raise HTTPException(
                    status_code=400, detail=f"Не удалось распознать роялти %: {royalty_percent!r}"
                )
            if not (0 <= value <= 100):
                raise HTTPException(
                    status_code=400,
                    detail=f"Роялти должно быть числом от 0 до 100, получено: {value}",
                )
            contragent.royalty_percent = value

    if reg_number is not None:
        if not reg_number.strip():
            contragent.reg_number = None
        else:
            type_for_check = contragent.type  # уже обновлён выше, если contragent_type передан
            value = normalize_reg_number(reg_number, type_for_check)
            conflict = (
                db.query(Contragent)
                .filter(Contragent.reg_number == value, Contragent.id != contragent.id)
                .one_or_none()
            )
            if conflict is not None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Рег. номер {value!r} уже занят контрагентом {conflict.title!r}",
                )
            contragent.reg_number = value

    if nicknames is not None:
        for old_nick in list(contragent.nicknames):
            db.delete(old_nick)
        db.flush()
        for nick in [n.strip() for n in nicknames.split(",") if n.strip()]:
            db.add(ContragentNickname(contragent_id=contragent.id, nickname=nick))

    # Реквизиты: переданный JSON ПОЛНОСТЬЮ заменяет прежний словарь (как
    # nicknames). Чистятся под текущий тип (contragent.type уже обновлён выше,
    # если тип меняли). Пустой после чистки -> None (реквизиты очищены).
    if requisites is not None:
        contragent.requisites = _parse_requisites(requisites, contragent.type) or None

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Рег. номер {reg_number!r} уже занят другим контрагентом",
        )

    log_action(
        db, current_user, "contragent.update", entity_type="contragent", entity_id=contragent.id,
        meta={"title": contragent.title},
    )

    return {
        "id": str(contragent.id),
        "title": contragent.title,
        "name": contragent.name,
        "country": contragent.country,
        "type": contragent.type,
        "contract_family": contragent.contract_family,
        "contract_date": (
            contragent.contract_date.isoformat() if contragent.contract_date else None
        ),
        "contract_number": contragent.contract_number,
        "royalty_percent": (
            float(contragent.royalty_percent) if contragent.royalty_percent is not None else None
        ),
        "reg_number": contragent.reg_number,
        "nicknames": [n.nickname for n in contragent.nicknames],
        "requisites": contragent.requisites or {},
    }


@contragents_router.delete("/{contragent_id}", dependencies=[Depends(require_role(ADMIN))])
def delete_contragent(
    contragent_id: uuid.UUID,
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Удаляет контрагента. Никнеймы удаляются каскадно (cascade="all,
    delete-orphan" в models.py + ondelete="CASCADE" на FK в БД).

    Пока нет таблицы generated_documents (см. брейншторм — сознательно не
    делаем), нет и проверки "а генерировали ли уже документы для этого
    контрагента" — удаление ничего, кроме самой карточки и её никнеймов,
    не затрагивает: ранее сгенерированные .docx в MinIO не трогаются,
    шаблоны и их теги тоже.

    Пока доступно только через Swagger — кнопка удаления в самом
    интерфейсе (экран документов контрагента) ещё не добавлена.
    """
    contragent = db.get(Contragent, contragent_id)
    if contragent is None:
        raise HTTPException(status_code=404, detail="Контрагент не найден")

    deleted_title = contragent.title  # запомнить до удаления — после db.delete() поле недоступно
    db.delete(contragent)
    db.commit()

    log_action(
        db, current_user, "contragent.delete", entity_type="contragent", entity_id=contragent_id,
        meta={"title": deleted_title},
    )

    return {"id": str(contragent_id), "deleted": True}


@contragents_router.get("/{contragent_id}/templates")
def list_contragent_templates(
    contragent_id: uuid.UUID,
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Подбор документов для контрагента по тегам.

    Обязательные теги подбора — country и contragent_type. Тип договора
    (contract_family) — УСЛОВНЫЙ фильтр, и работает ПО-РАЗНОМУ для договора и
    для приложения/акта (см. app/tags.py про две оси: платёж × обязательство):
      - если у контрагента он ЗАДАН:
          * ДОГОВОР — точное совпадение семейства (одно из 4);
          * ПРИЛОЖЕНИЕ/АКТ — совпадение по бакету обязательства: семейство
            контрагента сворачивается в 'ОБЯЗАТЕЛЬСТВО'/'БЕЗ_ОБЯЗАТЕЛЬСТВА', и
            берутся приложения/акты этого бакета (их платёжная ось не важна:
            акт роялти = акт аванс);
      - если НЕ задан — всё сразу (country + type): все 4 семейства договоров
        и оба бакета приложений/актов, оператор выбирает нужное.
    То есть заполненный тип договора сужает выдачу, пустой её не ограничивает
    (по просьбе пользователя).

    Надёжность подбора по-прежнему держится на нормализации регистра при
    создании контрагента (см. app/tags.py) и при тегировании шаблонов: если
    где-то закрадётся несовпадающий регистр, документ просто не найдётся без
    явной ошибки — держать в голове при отладке.

    Если у контрагента не заполнены country/type ("неполная" карточка из
    мягкого импорта) — возвращает пустой список, а не ошибку: карточка
    валидна, просто пока не участвует в подборе.

    contract_family возвращается у каждого шаблона: он нужен фронту, чтобы к
    открытому Приложению подобрать парный Акт. У приложения и акта это ОДИН И
    ТОТ ЖЕ бакет обязательства, поэтому фронт по-прежнему сравнивает их
    contract_family напрямую (см. DocFormPage) — просто теперь это бакет, а не
    полное семейство.

    Скрытые шаблоны (hidden_for_managers): менеджеру в подбор не попадают,
    остальным ролям (SEES_HIDDEN_TEMPLATES) — как обычно.
    """
    contragent = db.get(Contragent, contragent_id)
    if contragent is None:
        raise HTTPException(status_code=404, detail="Контрагент не найден")

    if not (contragent.country and contragent.type):
        return {"contragent_id": str(contragent_id), "templates": []}

    templates_query = db.query(Template).filter(
        Template.country == contragent.country,
        Template.contragent_type == contragent.type,
    )
    # Тип договора сужает выдачу ТОЛЬКО если задан — и по-разному для разных
    # типов документов (см. app/tags.py про две оси):
    #   - ДОГОВОР (и прочие типы) — точное совпадение семейства из 4 значений;
    #   - ПРИЛОЖЕНИЕ/АКТ — совпадение по бакету обязательства (у них платёжной
    #     оси нет): семейство контрагента сворачивается в бакет функцией
    #     obligation_bucket, а приложения/акты тегированы этим же бакетом.
    # Так одно «обычное приложение» обслуживает и аванс, и роялти, а
    # «приложение с обязательством» — и аванс+обязательство, и роялти+обязательство.
    if contragent.contract_family:
        bucket = obligation_bucket(contragent.contract_family)
        templates_query = templates_query.filter(
            or_(
                and_(
                    Template.doc_type.in_(OBLIGATION_DOC_TYPES),
                    Template.contract_family == bucket,
                ),
                and_(
                    or_(
                        Template.doc_type.notin_(OBLIGATION_DOC_TYPES),
                        Template.doc_type.is_(None),
                    ),
                    Template.contract_family == contragent.contract_family,
                ),
            )
        )
    if current_user.role not in SEES_HIDDEN_TEMPLATES:
        templates_query = templates_query.filter(Template.hidden_for_managers.is_(False))
    # По важности типа документа (договор→приложение→акт), внутри — по имени
    templates = templates_query.order_by(doc_type_sort_key(), Template.name).all()

    return {
        "contragent_id": str(contragent_id),
        "templates": [
            {
                "id": str(t.id),
                "name": t.name,
                "doc_type": t.doc_type,
                "contract_family": t.contract_family,
            }
            for t in templates
        ],
    }


@contragents_router.post(
    "/{contragent_id}/nicknames", dependencies=[Depends(require_role(*CAN_EDIT_CONTRAGENTS))]
)
def add_nickname(
    contragent_id: uuid.UUID,
    nickname: str = Form(...),
    db: Session = Depends(get_session),
) -> dict:
    """
    Добавляет псевдоним контрагенту. Никнеймов может быть несколько —
    на форме генерации они показываются выпадающим списком, не
    свободным текстом (см. брейншторм).
    """
    contragent = db.get(Contragent, contragent_id)
    if contragent is None:
        raise HTTPException(status_code=404, detail="Контрагент не найден")

    nick = ContragentNickname(contragent_id=contragent_id, nickname=nickname)
    db.add(nick)
    db.commit()

    return {"id": str(nick.id), "contragent_id": str(contragent_id), "nickname": nick.nickname}
